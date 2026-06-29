"""
2026年8月排班表生成脚本
基于门诊排班skill规则，从7月末状态衔接推算
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# 7月末衔接状态 (7/31 周五)
# 7/31(五): E=二院派遣, L=陈东升, 休=陈胜 (第5周备班=陈胜)
# 8月跨月自然周 7/27(一)-8/2(日)，备班保持陈胜不变
# 备班轮换顺序: 二院派遣(姜湛乾) → 陈东升 → 陈胜
# ============================================================

# ───── Week Cross (7/27-8/2): backup=陈胜 (跨月延续七月) ─────
# 7/31(五) E=二院派遣, L=陈东升, 休=陈胜 ✓
# 8/1(六): 周五晚(陈东升)→早, 周五休(陈胜)→晚 → E陈东升, L陈胜, 休二院派遣
# 8/2(日): 周六晚(陈胜)→早, 周六休(二院派遣)→晚 → E陈胜, L二院派遣, 休陈东升

# ───── Week 2 (8/3-8/9): backup=二院派遣 ─────
# Sun(8/2)晚=二院派遣 IS backup → 周日休班人(陈东升)做周一早
# 8/3(一): E=陈东升, L=陈胜, 备=二院派遣
# 8/4(二): E=陈胜, L=陈东升, 备=二院派遣
# 8/5(三): E=陈东升, L=陈胜, 备=二院派遣
# 8/6(四): E=陈胜, L=陈东升, 备=二院派遣
# 8/7(五): E=陈东升, L=陈胜, 休=二院派遣
# 8/8(六): 周五晚(陈胜)→早, 周五休(二院派遣)→晚 → E陈胜, L二院派遣, 休陈东升
# 8/9(日): 周六晚(二院派遣)→早, 周六休(陈东升)→晚 → E二院派遣, L陈东升, 休陈胜
# Sun晚=陈东升, 下周一备班=陈东升 → 周日休班人(陈胜)做周一早

# ───── Week 3 (8/10-8/16): backup=陈东升 ─────
# Sun(8/9)晚=陈东升 IS backup → 周日休班人(陈胜)做周一早
# 8/10(一): E=陈胜, L=二院派遣, 备=陈东升
# 8/11(二): E=二院派遣, L=陈胜, 备=陈东升
# 8/12(三): E=陈胜, L=二院派遣, 备=陈东升
# 8/13(四): E=二院派遣, L=陈胜, 备=陈东升
# 8/14(五): E=陈胜, L=二院派遣, 休=陈东升
# 8/15(六): 周五晚(二院派遣)→早, 周五休(陈东升)→晚 → E二院派遣, L陈东升, 休陈胜
# 8/16(日): 周六晚(陈东升)→早, 周六休(陈胜)→晚 → E陈东升, L陈胜, 休二院派遣
# Sun晚=陈胜, 下周一备班=陈胜 → 周日休班人(二院派遣)做周一早

# ───── Week 4 (8/17-8/23): backup=陈胜 ─────
# Sun(8/16)晚=陈胜 IS backup → 周日休班人(二院派遣)做周一早
# 8/17(一): E=二院派遣, L=陈东升, 备=陈胜
# 8/18(二): E=陈东升, L=二院派遣, 备=陈胜
# 8/19(三): E=二院派遣, L=陈东升, 备=陈胜
# 8/20(四): E=陈东升, L=二院派遣, 备=陈胜
# 8/21(五): E=二院派遣, L=陈东升, 休=陈胜
# 8/22(六): 周五晚(陈东升)→早, 周五休(陈胜)→晚 → E陈东升, L陈胜, 休二院派遣
# 8/23(日): 周六晚(陈胜)→早, 周六休(二院派遣)→晚 → E陈胜, L二院派遣, 休陈东升
# Sun晚=二院派遣, 下周一备班=二院派遣 → 周日休班人(陈东升)做周一早

# ───── Week 5 (8/24-8/30): backup=二院派遣 ─────
# Sun(8/23)晚=二院派遣 IS backup → 周日休班人(陈东升)做周一早
# 8/24(一): E=陈东升, L=陈胜, 备=二院派遣
# 8/25(二): E=陈胜, L=陈东升, 备=二院派遣
# 8/26(三): E=陈东升, L=陈胜, 备=二院派遣
# 8/27(四): E=陈胜, L=陈东升, 备=二院派遣
# 8/28(五): E=陈东升, L=陈胜, 休=二院派遣
# 8/29(六): 周五晚(陈胜)→早, 周五休(二院派遣)→晚 → E陈胜, L二院派遣, 休陈东升
# 8/30(日): 周六晚(二院派遣)→早, 周六休(陈东升)→晚 → E二院派遣, L陈东升, 休陈胜
# Sun晚=陈东升, 下周一备班=陈东升 → 周日休班人(陈胜)做周一早

# ───── Week 6 (8/31 only): backup=陈东升 ─────
# 8/31(一): 上周日晚(陈东升) IS backup → 周日休班人(陈胜)做周一早
# E=陈胜, L=二院派遣, 备=陈东升

# ============================================================
# Build schedule data
# ============================================================
sched = [
    ("W1", "第1周 (7/27-8/2)  |  备班人员：陈胜（跨月延续七月末班次）"),

    ("7/27", "一", "二院派遣", "陈东升", "陈胜", "", False),
    ("7/28", "二", "陈东升", "二院派遣", "陈胜", "", False),
    ("7/29", "三", "二院派遣", "陈东升", "陈胜", "", False),
    ("7/30", "四", "陈东升", "二院派遣", "陈胜", "", False),
    ("7/31", "五", "二院派遣", "陈东升", "", "陈胜", True),
    # 8/1(六): 周五晚(陈东升)→早, 周五休(陈胜)→晚
    ("8/1", "六", "陈东升", "陈胜", "", "二院派遣", True),
    # 8/2(日): 周六晚(陈胜)→早, 周六休(二院派遣)→晚
    ("8/2", "日", "陈胜", "二院派遣", "", "陈东升", True),

    ("W2", "第2周 (8/3-8/9)  |  备班人员：二院派遣"),

    # 8/3(一): Sun晚=二院派遣IS备班 → 周日休班(陈东升)做周一早
    ("8/3", "一", "陈东升", "陈胜", "二院派遣", "", False),
    ("8/4", "二", "陈胜", "陈东升", "二院派遣", "", False),
    ("8/5", "三", "陈东升", "陈胜", "二院派遣", "", False),
    ("8/6", "四", "陈胜", "陈东升", "二院派遣", "", False),
    ("8/7", "五", "陈东升", "陈胜", "", "二院派遣", True),
    # 8/8(六): 周五晚(陈胜)→早, 周五休(二院派遣)→晚
    ("8/8", "六", "陈胜", "二院派遣", "", "陈东升", True),
    # 8/9(日): 周六晚(二院派遣)→早, 周六休(陈东升)→晚
    ("8/9", "日", "二院派遣", "陈东升", "", "陈胜", True),

    ("W3", "第3周 (8/10-8/16)  |  备班人员：陈东升"),

    # 8/10(一): Sun晚=陈东升IS备班 → 周日休班(陈胜)做周一早
    ("8/10", "一", "陈胜", "二院派遣", "陈东升", "", False),
    ("8/11", "二", "二院派遣", "陈胜", "陈东升", "", False),
    ("8/12", "三", "陈胜", "二院派遣", "陈东升", "", False),
    ("8/13", "四", "二院派遣", "陈胜", "陈东升", "", False),
    ("8/14", "五", "陈胜", "二院派遣", "", "陈东升", True),
    # 8/15(六): 周五晚(二院派遣)→早, 周五休(陈东升)→晚
    ("8/15", "六", "二院派遣", "陈东升", "", "陈胜", True),
    # 8/16(日): 周六晚(陈东升)→早, 周六休(陈胜)→晚
    ("8/16", "日", "陈东升", "陈胜", "", "二院派遣", True),

    ("W4", "第4周 (8/17-8/23)  |  备班人员：陈胜"),

    # 8/17(一): Sun晚=陈胜IS备班 → 周日休班(二院派遣)做周一早
    ("8/17", "一", "二院派遣", "陈东升", "陈胜", "", False),
    ("8/18", "二", "陈东升", "二院派遣", "陈胜", "", False),
    ("8/19", "三", "二院派遣", "陈东升", "陈胜", "", False),
    ("8/20", "四", "陈东升", "二院派遣", "陈胜", "", False),
    ("8/21", "五", "二院派遣", "陈东升", "", "陈胜", True),
    # 8/22(六): 周五晚(陈东升)→早, 周五休(陈胜)→晚
    ("8/22", "六", "陈东升", "陈胜", "", "二院派遣", True),
    # 8/23(日): 周六晚(陈胜)→早, 周六休(二院派遣)→晚
    ("8/23", "日", "陈胜", "二院派遣", "", "陈东升", True),

    ("W5", "第5周 (8/24-8/30)  |  备班人员：二院派遣"),

    # 8/24(一): Sun晚=二院派遣IS备班 → 周日休班(陈东升)做周一早
    ("8/24", "一", "陈东升", "陈胜", "二院派遣", "", False),
    ("8/25", "二", "陈胜", "陈东升", "二院派遣", "", False),
    ("8/26", "三", "陈东升", "陈胜", "二院派遣", "", False),
    ("8/27", "四", "陈胜", "陈东升", "二院派遣", "", False),
    ("8/28", "五", "陈东升", "陈胜", "", "二院派遣", True),
    # 8/29(六): 周五晚(陈胜)→早, 周五休(二院派遣)→晚
    ("8/29", "六", "陈胜", "二院派遣", "", "陈东升", True),
    # 8/30(日): 周六晚(二院派遣)→早, 周六休(陈东升)→晚
    ("8/30", "日", "二院派遣", "陈东升", "", "陈胜", True),

    ("W6", "第6周 (8/31-8/31)  |  备班人员：陈东升（仅周一一天）"),

    # 8/31(一): Sun晚=陈东升IS备班 → 周日休班(陈胜)做周一早
    ("8/31", "一", "陈胜", "二院派遣", "陈东升", "", False),
]

# ============================================================
# Styles
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = "排班表"

hf = Font(name="Arial", bold=True, size=11, color="FFFFFF")
hfl = PatternFill("solid", fgColor="2C2C2A")
wlf = PatternFill("solid", fgColor="D3D1C7")
dfl = PatternFill("solid", fgColor="F1EFE8")
wf = PatternFill("solid", fgColor="FAEEDA")
wdf = PatternFill("solid", fgColor="FAC775")
ef = Font(name="Arial", bold=True, size=11, color="0C447C")
lf = Font(name="Arial", bold=True, size=11, color="712B13")
bf = Font(name="Arial", bold=True, size=11, color="085041")
offf = Font(name="Arial", bold=True, size=11, color="A32D2D")
df = Font(name="Arial", bold=True, size=11, color="888780")
lab = Font(name="Arial", bold=True, size=11, color="2C2C2A")
nf = Font(name="Arial", size=11, color="444441")
th = Border(
    left=Side(style="thin", color="B4B2A9"),
    right=Side(style="thin", color="B4B2A9"),
    top=Side(style="thin", color="B4B2A9"),
    bottom=Side(style="thin", color="B4B2A9"),
)
ca = Alignment(horizontal="center", vertical="center")
la = Alignment(horizontal="left", vertical="center")

# === Title ===
ws.merge_cells("A1:F1")
ws["A1"] = "2026年8月工作排班表（8月1日 — 8月31日）"
ws["A1"].font = Font(name="Arial", bold=True, size=14, color="2C2C2A")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

# === Headers ===
headers = ["日期", "星期", "早班\n08:00-14:00", "晚班\n14:00-20:00", "备班\n08:00-12:00\n+14:00-17:30", "休假"]
ws.row_dimensions[3].height = 40
for i, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=i, value=h)
    c.font = hf
    c.fill = hfl
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = th

# === Write schedule ===
row = 4
for item in sched:
    if isinstance(item[0], str) and item[0].startswith("W"):
        # Week header row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row=row, column=1, value=item[1])
        c.font = lab
        c.fill = wlf
        c.alignment = la
        c.border = th
        for cc in range(2, 7):
            ws.cell(row=row, column=cc).fill = wlf
            ws.cell(row=row, column=cc).border = th
        ws.row_dimensions[row].height = 24
        row += 1
    else:
        date, wd, early, late, backup, off_p, is_we = item
        data = [date, wd, early, late, backup, off_p if off_p else "-"]
        for c_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.border = th
            cell.alignment = ca
            if c_idx == 1:
                cell.fill = wdf if is_we else dfl
                cell.font = lab
            elif c_idx == 2:
                cell.fill = wf if is_we else dfl
                cell.font = lab
            else:
                if is_we:
                    cell.fill = wf
                # 派遣人员标红底黑字
                rfl = PatternFill("solid", fgColor="FF6666")
                if val == "二院派遣":
                    cell.font = nf
                    cell.fill = rfl
                elif c_idx == 3:
                    cell.font = ef  # 早班=蓝色
                elif c_idx == 4:
                    cell.font = lf  # 晚班=橙色
                elif c_idx == 5:
                    cell.font = df if val == "-" else bf  # 备班=绿色
                elif c_idx == 6:
                    cell.font = df if val == "-" else offf  # 休假=红色
        ws.row_dimensions[row].height = 24
        row += 1

# === Column widths ===
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 8
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 24
ws.column_dimensions["F"].width = 42

fp = "/home/bobobears/下载/排班表_2026年8月.xlsx"
wb.save(fp)
print(f"✅ 已生成：{fp}")

# ============================================================
# Audit: triple verification
# ============================================================
print("\n═══════════ 审计验证 ═══════════")

errors = []
# Only audit August dates
aug_items = [it for it in sched if not (isinstance(it[0], str) and it[0].startswith("W")) and it[0].startswith("8/")]

# Audit 1: 当日早班≠晚班
for it in aug_items:
    date, wd, early, late, backup, off_p, is_we = it
    if early == late:
        errors.append(f"❌ [{date}] 早班=晚班={early}")

# Audit 2: 连续两天早/晚班不重复
for i in range(1, len(aug_items)):
    prev = aug_items[i-1]
    curr = aug_items[i]
    if prev[2] == curr[2]:  # same early person consecutive days
        errors.append(f"❌ [{prev[0]}→{curr[0]}] 连续早班={prev[2]}")
    if prev[3] == curr[3]:  # same late person consecutive days
        errors.append(f"❌ [{prev[0]}→{curr[0]}] 连续晚班={prev[3]}")

# Audit 3: 休假顺序
for it in aug_items:
    date, wd, early, late, backup, off_p, is_we = it
    if wd == "五" and off_p:
        # Friday: backup person should be off
        pass  # We'd need to know the week's backup, checked below
    if wd == "六":
        # Find previous Friday
        pass

# Detailed weekend check
# Group by week label
week_groups = {}
current_wlabel = None
for it in sched:
    if isinstance(it[0], str) and it[0].startswith("W"):
        current_wlabel = it[1]
        week_groups[current_wlabel] = []
    elif it[0].startswith("8/"):
        week_groups[current_wlabel].append(it)

for wdesc, items in week_groups.items():
    # Find Fri, Sat, Sun
    fri_item = None
    sat_item = None
    sun_item = None
    for it in items:
        if it[1] == "五":
            fri_item = it
        elif it[1] == "六":
            sat_item = it
        elif it[1] == "日":
            sun_item = it

    import re
    m = re.search(r'备班人员[：:]\s*([^（(]+)', wdesc)
    backup_person = m.group(1).strip() if m else "?"

    if fri_item:
        if fri_item[5] != backup_person:
            errors.append(f"❌ [{fri_item[0]}(五)] 休假={fri_item[5]}，应休={backup_person}（备班人）")
    if fri_item and sat_item:
        # Sat off = Fri early
        expected_sat_off = fri_item[2]  # Fri early
        if sat_item[5] != expected_sat_off:
            errors.append(f"❌ [{sat_item[0]}(六)] 休假={sat_item[5]}，应休={expected_sat_off}（周五早班人）")
    if fri_item and sun_item:
        # Sun off = Fri late
        expected_sun_off = fri_item[3]  # Fri late
        if sun_item[5] != expected_sun_off:
            errors.append(f"❌ [{sun_item[0]}(日)] 休假={sun_item[5]}，应休={expected_sun_off}（周五晚班人）")

if errors:
    print(f"\n🔴 发现 {len(errors)} 个问题：")
    for e in errors:
        print(f"  {e}")
else:
    print("✅ 审计全部通过！")
    print("  1. 当日早班≠晚班 ✓")
    print("  2. 连续两天早/晚班不重复 ✓")
    print("  3. 休假顺序正确 ✓")

print(f"\n📊 排班统计：")
print(f"  已生成排班表，详细见 Excel 文件")
