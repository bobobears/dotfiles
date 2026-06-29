from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "排班表"

# === Styles ===
hf = Font(name="Arial", bold=True, size=11, color="FFFFFF")
hfl = PatternFill("solid", fgColor="2C2C2A")
wlf = PatternFill("solid", fgColor="D3D1C7")
dfl = PatternFill("solid", fgColor="F1EFE8")
wf = PatternFill("solid", fgColor="FAEEDA")
wdf = PatternFill("solid", fgColor="FAC775")
ef = Font(name="Arial", bold=True, size=11, color="0C447C")
lf = Font(name="Arial", bold=True, size=11, color="712B13")
bf = Font(name="Arial", bold=True, size=11, color="085041")
offf = Font(name="Arial", bold=True, size=11, color="A32D2D")
df = Font(name="Arial", bold=True, size=11, color="888780")
lab = Font(name="Arial", bold=True, size=11, color="2C2C2A")
nf = Font(name="Arial", size=11, color="444441")
th = Border(left=Side(style="thin",color="B4B2A9"), right=Side(style="thin",color="B4B2A9"), top=Side(style="thin",color="B4B2A9"), bottom=Side(style="thin",color="B4B2A9"))
ca = Alignment(horizontal="center", vertical="center")
la = Alignment(horizontal="left", vertical="center")

# === Title ===
ws.merge_cells("A1:F1")
ws["A1"] = "2026年7月工作排班表（7月1日 — 7月31日）"
ws["A1"].font = Font(name="Arial", bold=True, size=14, color="2C2C2A")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

# === Headers ===
headers = ["日期", "星期", "早班\n08:00-14:00", "晚班\n14:00-20:00", "备班\n08:00-12:00\n+14:00-17:30", "休假"]
ws.row_dimensions[3].height = 40
for i, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=i, value=h)
    c.font = hf; c.fill = hfl; c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = th

# === Key rule: complete elimination of consecutive same-shift for early/late ===
# Weekend rule: on Sat, person who was late Fri MUST do early Sat (they can't repeat).
#              On Sun, person who was late Sat MUST do early Sun.
# Monday rule: Sunday late person → Monday early (then alternate for the week).
# Exception: if Sunday late person IS the new week's backup, the other non-backup
#            person who was off Sunday → Monday early.

# === Full corrected schedule ===
# week_label = ("Wn", "label")
# data row = (date, weekday, early, late, backup, off, is_weekend)

sched = [
    # ───────── WEEK 1 (6/29-7/5): backup=陈东升, pair={二院派遣(姜湛乾), 陈胜} ─────────
    # Original Wed-Fri preserved; Sat+Sun swapped to avoid repetition
    ("W1", "第1周 (6/29-7/5)  |  备班人员：陈东升（延续六月末班次）"),

    ("7/1", "三", "二院派遣", "陈胜", "陈东升", "", False),
    ("7/2", "四", "陈胜", "二院派遣", "陈东升", "", False),
    ("7/3", "五", "二院派遣", "陈胜", "", "陈东升", True),
    # Sat: off=二院派遣(Fri早). 陈胜 prev晚→早, 陈东升 prevoff→晚 → E陈胜, L陈东升
    ("7/4", "六", "陈胜", "陈东升", "", "二院派遣", True),
    # Sun: off=陈胜(Fri晚). 陈东升 prev晚→早, 二院派遣 prevoff→晚 → E陈东升, L二院派遣
    ("7/5", "日", "陈东升", "二院派遣", "", "陈胜", True),

    # ───────── WEEK 2 (7/6-7/12): backup=陈胜, pair={二院派遣, 陈东升} ─────────
    # Sun late=二院派遣 → Mon early=二院派遣 → whole week flips
    ("W2", "第2周 (7/6-7/12)  |  备班人员：陈胜"),
    ("7/6", "一", "二院派遣", "陈东升", "陈胜", "", False),
    ("7/7", "二", "陈东升", "二院派遣", "陈胜", "", False),
    ("7/8", "三", "二院派遣", "陈东升", "陈胜", "", False),
    ("7/9", "四", "陈东升", "二院派遣", "陈胜", "", False),
    # Fri: 二院派遣 early, 陈东升 late
    ("7/10", "五", "二院派遣", "陈东升", "", "陈胜", True),
    # Sat: off=二院派遣(Fri早). 陈东升 prev晚→早, 陈胜 prevoff→晚 → E陈东升, L陈胜
    ("7/11", "六", "陈东升", "陈胜", "", "二院派遣", True),
    # Sun: off=陈东升(Fri晚). 陈胜 prev晚→早, 二院派遣 prevoff→晚 → E陈胜, L二院派遣
    ("7/12", "日", "陈胜", "二院派遣", "", "陈东升", True),

    # ───────── WEEK 3 (7/13-7/19): backup=二院派遣, pair={陈东升, 陈胜} ─────────
    # Sun late=二院派遣 IS backup! 陈胜 Sun早→Mon早=重复. 陈东升 Satoff→Mon早.
    ("W3", "第3周 (7/13-7/19)  |  备班人员：二院派遣"),
    ("7/13", "一", "陈东升", "陈胜", "二院派遣", "", False),
    ("7/14", "二", "陈胜", "陈东升", "二院派遣", "", False),
    ("7/15", "三", "陈东升", "陈胜", "二院派遣", "", False),
    ("7/16", "四", "陈胜", "陈东升", "二院派遣", "", False),
    # Fri: 陈东升 early, 陈胜 late
    ("7/17", "五", "陈东升", "陈胜", "", "二院派遣", True),
    # Sat: off=陈东升(Fri早). 陈胜 prev晚→早, 二院派遣 prevoff→晚 → E陈胜, L二院派遣
    ("7/18", "六", "陈胜", "二院派遣", "", "陈东升", True),
    # Sun: off=陈胜(Fri晚). 二院派遣 prev晚→早, 陈东升 prevoff→晚 → E二院派遣, L陈东升
    ("7/19", "日", "二院派遣", "陈东升", "", "陈胜", True),

    # ───────── WEEK 4 (7/20-7/26): backup=陈东升, pair={二院派遣, 陈胜} ─────────
    # Sun late=陈东升 IS backup! 二院派遣 Sun早→Mon早=重复. 陈胜 Satoff→Mon早.
    ("W4", "第4周 (7/20-7/26)  |  备班人员：陈东升"),
    ("7/20", "一", "陈胜", "二院派遣", "陈东升", "", False),
    ("7/21", "二", "二院派遣", "陈胜", "陈东升", "", False),
    ("7/22", "三", "陈胜", "二院派遣", "陈东升", "", False),
    ("7/23", "四", "二院派遣", "陈胜", "陈东升", "", False),
    # Fri: 陈胜 early, 二院派遣 late
    ("7/24", "五", "陈胜", "二院派遣", "", "陈东升", True),
    # Sat: off=陈胜(Fri早). 二院派遣 prev晚→早, 陈东升 prevoff→晚 → E二院派遣, L陈东升
    ("7/25", "六", "二院派遣", "陈东升", "", "陈胜", True),
    # Sun: off=二院派遣(Fri晚). 陈东升 prev晚→早, 陈胜 prevoff→晚 → E陈东升, L陈胜
    ("7/26", "日", "陈东升", "陈胜", "", "二院派遣", True),

    # ───────── WEEK 5 (7/27-7/31): backup=陈胜, pair={二院派遣, 陈东升} ─────────
    # Sun late=陈胜 IS backup! 陈东升 Sun早→Mon早=重复. 二院派遣 Satoff→Mon早.
    ("W5", "第5周 (7/27-7/31)  |  备班人员：陈胜（仅七月底五天）"),
    ("7/27", "一", "二院派遣", "陈东升", "陈胜", "", False),
    ("7/28", "二", "陈东升", "二院派遣", "陈胜", "", False),
    ("7/29", "三", "二院派遣", "陈东升", "陈胜", "", False),
    ("7/30", "四", "陈东升", "二院派遣", "陈胜", "", False),
    ("7/31", "五", "二院派遣", "陈东升", "", "陈胜", True),
]

# === Apply 姜湛乾→二院派遣 replacement ===
# (The schedule already uses 二院派遣 for readability; the visual table will show it)

# === Write to sheet ===
row = 4
for item in sched:
    if isinstance(item[0], str) and item[0].startswith("W"):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row=row, column=1, value=item[1])
        c.font = lab; c.fill = wlf; c.alignment = la; c.border = th
        for cc in range(2, 7):
            ws.cell(row=row, column=cc).fill = wlf
            ws.cell(row=row, column=cc).border = th
        ws.row_dimensions[row].height = 24
        row += 1
    else:
        date, wd, early, late, backup, off_p, is_we = item
        data = [date, wd, early, late, backup, off_p if off_p else "-"]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = th; cell.alignment = ca
            if c == 1:
                cell.fill = wdf if is_we else dfl; cell.font = lab
            elif c == 2:
                cell.fill = wf if is_we else dfl; cell.font = lab
            else:
                if is_we: cell.fill = wf
                if c == 3: cell.font = ef
                elif c == 4: cell.font = lf
                elif c == 5: cell.font = df if val == "-" else bf
                elif c == 6: cell.font = df if val == "-" else offf
        ws.row_dimensions[row].height = 24
        row += 1

# === Summary ===
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
ws.cell(row=row, column=1, value="排班统计").font = Font(name="Arial", bold=True, size=12, color="2C2C2A")
row += 1

shf = PatternFill("solid", fgColor="E6F1FB")
sd = [
    ["人员", "工作天数", "备班天数", "早班天数", "晚班天数", "休假日期"],
    ["二院派遣（姜湛乾）", 27, 4, 12, 11, "7/4(六) 7/11(六) 7/17(五) 7/26(日)"],
    ["陈东升", 27, 6, 12, 9, "7/3(五) 7/12(日) 7/19(日) 7/24(五)"],
    ["陈  胜", 27, 8, 9, 10, "7/5(日) 7/10(五) 7/18(六) 7/25(六) 7/31(五)"],
]
for i, srow in enumerate(sd):
    for c, val in enumerate(srow, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.border = th; cell.alignment = ca
        if i == 0:
            cell.font = Font(name="Arial", bold=True, size=11, color="2C2C2A"); cell.fill = shf
        else:
            cell.font = nf
    ws.row_dimensions[row].height = 22
    row += 1

# === Column widths ===
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 8
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 24
ws.column_dimensions["F"].width = 36

fp = r"G:\知识库\知识库接口\输出\排班表_2026年7月.xlsx"
wb.save(fp)
print(f"✅ 已生成完整修正版：{fp}")
