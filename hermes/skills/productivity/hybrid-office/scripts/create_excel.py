#!/usr/bin/env python3
"""
Create a professionally styled Excel workbook (.xlsx).

Usage:
    python create_excel.py <output_path>

Reads workbook structure from stdin as JSON:
{
  "title": "Report Title",
  "sheets": [
    {
      "name": "Sheet1",
      "freeze_pane": "A2",
      "column_widths": [12, 20, 15, 15],
      "tables": [
        {
          "headers": ["Column A", "Column B", "Column C"],
          "rows": [
            ["Data1", "Data2", "Data3"]
          ],
          "start_row": 1
        }
      ],
      "title_cell": {"row": 1, "col": 1, "text": "Report Title", "merge_to": {"row": 1, "col": 4}},
      "summary_rows": [
        {"row_type": "after_table", "text": "Total:", "bold": true}
      ]
    }
  ]
}
"""

import json
import sys
import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ── Professional Color Palette ──
DARK_BLUE_HEX = "1B3A5C"
ACCENT_BLUE_HEX = "2E6BA4"
LIGHT_BG_HEX = "F2F6FA"
WHITE_HEX = "FFFFFF"
BODY_FONT_HEX = "333333"
MUTED_HEX = "666666"
BORDER_COLOR = "BBBBBB"

DARK_BLUE = Font(color=DARK_BLUE_HEX, bold=True, size=18, name="Calibri")
ACCENT_FILL = PatternFill(start_color=ACCENT_BLUE_HEX, end_color=ACCENT_BLUE_HEX, fill_type="solid")
WHITE_BOLD = Font(color=WHITE_HEX, bold=True, size=10, name="Calibri")
BODY_FONT = Font(color=BODY_FONT_HEX, size=10, name="Calibri")
LIGHT_FILL = PatternFill(start_color=LIGHT_BG_HEX, end_color=LIGHT_BG_HEX, fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color=BORDER_COLOR),
    right=Side(style='thin', color=BORDER_COLOR),
    top=Side(style='thin', color=BORDER_COLOR),
    bottom=Side(style='thin', color=BORDER_COLOR),
)
BOTTOM_ACCENT = Border(
    bottom=Side(style='medium', color=ACCENT_BLUE_HEX),
    left=Side(style='thin', color=BORDER_COLOR),
    right=Side(style='thin', color=BORDER_COLOR),
    top=Side(style='thin', color=BORDER_COLOR),
)

# Table style
TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)


def auto_width(ws, column_widths=None, min_width=8, max_width=40):
    """Auto-set column widths based on content or explicit widths."""
    if column_widths:
        for i, w in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return

    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val.encode('utf-8')))
            except:
                max_len = max(max_len, len(str(cell.value or "")))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def build_workbook(structure, output_path):
    """Build a complete Excel workbook from a structure dict."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for sheet_data in structure.get("sheets", []):
        ws = wb.create_sheet(title=sheet_data.get("name", "Sheet"))

        # Default row height
        ws.sheet_properties.defaultRowHeight = 20

        # Column widths
        auto_width(ws, sheet_data.get("column_widths"))

        # Freeze pane
        freeze = sheet_data.get("freeze_pane")
        if freeze:
            ws.freeze_panes = freeze

        # Title cell (merged)
        title_cell = sheet_data.get("title_cell")
        if title_cell:
            r, c = title_cell["row"], title_cell["col"]
            cell = ws.cell(row=r, column=c, value=title_cell.get("text", ""))
            cell.font = DARK_BLUE
            cell.alignment = Alignment(horizontal="left", vertical="center")
            merge_to = title_cell.get("merge_to")
            if merge_to:
                ws.merge_cells(
                    start_row=r, start_column=c,
                    end_row=merge_to["row"], end_column=merge_to["col"]
                )
            ws.row_dimensions[r].height = 30

        # Tables
        for tbl_data in sheet_data.get("tables", []):
            headers = tbl_data.get("headers", [])
            rows = tbl_data.get("rows", [])
            start_row = tbl_data.get("start_row", 1)

            # Account for title row offset
            data_start = start_row

            # Write headers
            for c_idx, header_text in enumerate(headers, 1):
                cell = ws.cell(row=data_start, column=c_idx, value=header_text)
                cell.font = WHITE_BOLD
                cell.fill = ACCENT_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = THIN_BORDER

            # Write data
            for r_idx, row_data in enumerate(rows, 1):
                for c_idx, cell_text in enumerate(row_data, 1):
                    cell = ws.cell(row=data_start + r_idx, column=c_idx, value=cell_text)
                    cell.font = BODY_FONT
                    cell.alignment = Alignment(vertical="center")
                    cell.border = THIN_BORDER
                    # Alternating rows
                    if r_idx % 2 == 0:
                        cell.fill = LIGHT_FILL

            # Apply bottom accent border on header
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=data_start, column=c_idx).border = BOTTOM_ACCENT

            # Add as an Excel table for filtering
            if headers:
                last_row = data_start + len(rows)
                last_col = len(headers)
                table_ref = f"A{data_start}:{get_column_letter(last_col)}{last_row}"
                table = Table(displayName=f"Table{ws.title.replace(' ','')}",
                              ref=table_ref)
                table.tableStyleInfo = TABLE_STYLE
                ws.add_table(table)

            # Summary rows
            summary_rows = sheet_data.get("summary_rows", [])
            for s_row in summary_rows:
                rtype = s_row.get("row_type", "after_table")
                if rtype == "after_table":
                    sr = data_start + len(rows) + 1
                    text = s_row.get("text", "")
                    cols_count = len(headers) if headers else 1
                    cell = ws.cell(row=sr, column=1, value=text)
                    if s_row.get("bold"):
                        cell.font = Font(color=BODY_FONT_HEX, bold=True, size=10, name="Calibri")
                    else:
                        cell.font = BODY_FONT
                    cell.alignment = Alignment(vertical="center")
                    cell.border = THIN_BORDER
                    # Add top border for summary
                    cell.border = Border(
                        top=Side(style='thin', color=ACCENT_BLUE_HEX),
                        bottom=Side(style='thin', color=BORDER_COLOR),
                        left=Side(style='thin', color=BORDER_COLOR),
                        right=Side(style='thin', color=BORDER_COLOR),
                    )
                    # Merge across if needed
                    if cols_count > 1:
                        ws.merge_cells(start_row=sr, start_column=1,
                                       end_row=sr, end_column=cols_count)

        # Print settings
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = None
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5

    # Remove default extra sheet if any remain
    wb.save(output_path)
    print(f"✅ Excel workbook saved: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Create a professional Excel workbook")
    parser.add_argument("output", help="Output .xlsx file path")
    args = parser.parse_args()

    raw = sys.stdin.read()
    if not raw.strip():
        print("❌ No input JSON provided via stdin", file=sys.stderr)
        sys.exit(1)

    structure = json.loads(raw)
    build_workbook(structure, args.output)


if __name__ == "__main__":
    main()
