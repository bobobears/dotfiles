#!/usr/bin/env python3
"""
Edit an existing Excel workbook (.xlsx).

Usage:
    python edit_excel.py <input_path> <output_path> < edits.json

Edits JSON format:
{
  "changes": [
    {"action": "update_cell", "sheet": "Sheet1", "row": 2, "col": 1, "value": "New Value"},
    {"action": "update_cell_by_header", "sheet": "Sheet1", "header": "Name", "row": 3, "value": "John"},
    {"action": "add_row", "sheet": "Sheet1", "data": ["A", "B", "C"], "after_row": 5},
    {"action": "add_column", "sheet": "Sheet1", "header": "New Col", "default_value": ""},
    {"action": "rename_sheet", "old": "Sheet1", "new": "Data"},
    {"action": "add_sheet", "name": "Summary"},
    {"action": "set_cell_style", "sheet": "Sheet1", "row": 1, "col": 1, "bold": true, "color": "1B3A5C"},
    {"action": "delete_row", "sheet": "Sheet1", "row": 3}
  ]
}
"""

import json
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ACCENT_BLUE = "2E6BA4"
WHITE = "FFFFFF"
LIGHT_BG = "F2F6FA"
BODY_COLOR = "333333"
BORDER_COLOR = "BBBBBB"

DARK_BLUE_FONT = Font(color="1B3A5C", bold=True, size=18, name="Calibri")
WHITE_BOLD = Font(color=WHITE, bold=True, size=10, name="Calibri")
BODY_FONT = Font(color=BODY_COLOR, size=10, name="Calibri")
ACCENT_FILL = PatternFill(start_color=ACCENT_BLUE, end_color=ACCENT_BLUE, fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color=BORDER_COLOR),
    right=Side(style='thin', color=BORDER_COLOR),
    top=Side(style='thin', color=BORDER_COLOR),
    bottom=Side(style='thin', color=BORDER_COLOR),
)


def find_header_col(ws, header_name):
    """Find column index (1-based) of a header."""
    for cell in ws[1]:
        if cell.value and str(cell.value).strip() == header_name:
            return cell.column
    return None


def edit_excel(input_path, output_path, edits):
    """Apply edits to an Excel workbook."""
    wb = load_workbook(input_path)

    for edit in edits:
        action = edit.get("action")

        if action == "update_cell":
            ws = wb[edit.get("sheet", wb.sheetnames[0])]
            row, col = edit["row"], edit["col"]
            cell = ws.cell(row=row, column=col, value=edit["value"])
            cell.font = BODY_FONT

        elif action == "update_cell_by_header":
            ws = wb[edit.get("sheet", wb.sheetnames[0])]
            col = find_header_col(ws, edit["header"])
            if col:
                cell = ws.cell(row=edit["row"], column=col, value=edit["value"])
                cell.font = BODY_FONT

        elif action == "add_row":
            ws = wb[edit.get("sheet", wb.sheetnames[0])]
            data = edit.get("data", [])
            after_row = edit.get("after_row", ws.max_row)
            ws.insert_rows(after_row + 1)
            for c_idx, val in enumerate(data, 1):
                cell = ws.cell(row=after_row + 1, column=c_idx, value=val)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                if (after_row) % 2 == 1:
                    cell.fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")

        elif action == "add_column":
            ws = wb[edit.get("sheet", wb.sheetnames[0])]
            new_col = ws.max_column + 1
            header = edit.get("header", f"Column{new_col}")
            cell = ws.cell(row=1, column=new_col, value=header)
            cell.font = WHITE_BOLD
            cell.fill = ACCENT_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            default_val = edit.get("default_value", "")
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=new_col, value=default_val)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER

        elif action == "rename_sheet":
            old = edit.get("old")
            new = edit.get("new")
            if old in wb.sheetnames:
                wb[old].title = new

        elif action == "add_sheet":
            wb.create_sheet(title=edit.get("name", "Sheet"))

        elif action == "set_cell_style":
            ws = wb[edit.get("sheet", wb.sheetnames[0])]
            cell = ws.cell(row=edit["row"], column=edit["col"])
            if edit.get("bold"):
                cell.font = Font(color=edit.get("color", BODY_COLOR), bold=True, size=10, name="Calibri")

        elif action == "delete_row":
            ws = wb[edit.get("sheet", wb.sheetnames[0])]
            ws.delete_rows(edit["row"])

    wb.save(output_path)
    print(f"✅ Excel workbook edited and saved: {output_path}")
    return output_path


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Edit an existing Excel workbook")
    parser.add_argument("input", help="Input .xlsx file path")
    parser.add_argument("output", help="Output .xlsx file path")
    args = parser.parse_args()

    raw = sys.stdin.read()
    if not raw.strip():
        print("❌ No edit JSON provided via stdin", file=sys.stderr)
        sys.exit(1)

    edits_data = json.loads(raw)
    if isinstance(edits_data, dict) and "changes" in edits_data:
        edits = edits_data["changes"]
    else:
        edits = edits_data

    edit_excel(args.input, args.output, edits)


if __name__ == "__main__":
    main()
